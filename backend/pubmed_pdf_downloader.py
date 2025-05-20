import os
import time
import glob
import shutil
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class PubMedProcessor:
    def __init__(self, input_path, output_path):
        self.input_path = input_path
        self.output_path = output_path
        self.download_dir = os.path.join(os.getcwd(), "downloads")
        os.makedirs(self.download_dir, exist_ok=True)
        self.df = None
        self.driver = None
        self.wait = None

    def _setup_browser(self):
        chrome_options = Options()
        chrome_options.add_experimental_option("prefs", {
            "download.default_directory": self.download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        })
        chrome_options.add_argument("--start-maximized")
        self.driver = webdriver.Chrome(options=chrome_options)
        self.wait = WebDriverWait(self.driver, 20)

    def _read_excel(self):
        try:
            self.df = pd.read_excel(self.input_path)
            self.df['Comment'] = self.df.get('Comment', '')
            self.df['Result Count'] = self.df.get('Result Count', '')
        except Exception as e:
            raise Exception(f"Error reading Excel: {e}")

    def _save_and_style_excel(self):
        self.df.to_excel(self.output_path, index=False)
        try:
            wb = load_workbook(self.output_path)
            ws = wb.active

            # Style headers
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_num, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                for cell in col_cells:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    ws.column_dimensions[get_column_letter(col_num)].width = max(len(str(cell.value)) + 5, 15)

            # Align data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            wb.save(self.output_path)
        except Exception as e:
            raise Exception(f"Error styling Excel: {e}")

    def _process_search(self, index, query, filename):
        self.driver.get("https://pubmed.ncbi.nlm.nih.gov/")

        try:
            search_box = self.wait.until(EC.presence_of_element_located((By.ID, "id_term")))
            search_box.clear()
            search_box.send_keys(query)
            search_box.send_keys(Keys.ENTER)

            self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, "results-amount")))
            time.sleep(2)
            result_text = self.driver.find_element(By.CLASS_NAME, "results-amount").text
            result_count = int(result_text.split()[0].replace(",", ""))
            self.df.at[index, 'Result Count'] = result_count

            if result_count > 1000:
                self.df.at[index, 'Comment'] = 'Too many results'
                return

            for warning in self.driver.find_elements(By.CLASS_NAME, "usa-alert-body"):
                if "Quoted phrase not found in" in warning.text:
                    self.df.at[index, 'Comment'] = 'Quoted phrase warning'
                    raise Exception("Quoted phrase issue")

        except Exception:
            if not self.df.at[index, 'Comment']:
                self.df.at[index, 'Comment'] = 'Search/warning error'
            raise

        try:
            save_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Save')]")))
            save_btn.click()
            self.wait.until(EC.visibility_of_element_located((By.ID, "save-action-selection")))

            self.driver.find_element(By.ID, "save-action-selection").send_keys(Keys.DOWN + Keys.RETURN)
            time.sleep(0.5)
            self.driver.find_element(By.ID, "save-action-format").send_keys(Keys.DOWN + Keys.RETURN)
            time.sleep(0.5)

            create_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Create file')]")))
            create_btn.click()

            time.sleep(7)
            downloaded_files = glob.glob(os.path.join(self.download_dir, "pubmed-*.*"))

            if downloaded_files:
                latest_file = max(downloaded_files, key=os.path.getctime)
                new_path = os.path.join(self.download_dir, f"{filename}.txt")
                shutil.move(latest_file, new_path)
                self.df.at[index, 'Comment'] = 'Downloaded'
            else:
                self.df.at[index, 'Comment'] = 'Download error'
                raise Exception("No file downloaded")

        except Exception:
            self.df.at[index, 'Comment'] = 'Download error'
            raise

    def process(self):
        self._read_excel()
        self._setup_browser()

        try:
            for index, row in self.df.iterrows():
                query = row.get('Search Strategy 1')
                filename = row.get('File Name')

                if pd.isna(query) or pd.isna(filename):
                    self.df.at[index, 'Comment'] = 'Empty values'
                    continue

                print(f"Searching for: {query}")
                try:
                    self._process_search(index, query, filename)
                except Exception as e:
                    print(f"Row {index + 2} failed: {e}")
                    continue
        finally:
            self.driver.quit()

        self._save_and_style_excel()
